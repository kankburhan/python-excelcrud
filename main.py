from flask import Flask, render_template, request, redirect, url_for, jsonify
from openpyxl import load_workbook
from datetime import datetime

app = Flask(__name__)

# Path to the Excel file
EXCEL_FILE = 'report.xlsx'

# Function to get the next ID for a new row
def get_next_id(sheet_name):
    wb = load_workbook(EXCEL_FILE)
    ws = wb[sheet_name]
    max_id = 0
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
        product_id = row[0]  # Assuming ID is in the first column
        if product_id and isinstance(product_id, int):
            max_id = max(max_id, product_id)
    return max_id + 1  # Increment ID by 1

# Read data from the Excel file
def read_excel(sheet_name):
    wb = load_workbook(EXCEL_FILE)
    ws = wb[sheet_name]
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
        data.append(list(row))
    return data

# Write data to the Excel file
def write_to_excel(sheet_name, row_data):
    wb = load_workbook(EXCEL_FILE)
    ws = wb[sheet_name]
    
    # Get the next ID
    next_id = get_next_id(sheet_name)
    
    # Prepend the ID to the row data
    row_data.insert(0, next_id)
    
    # Append the current date and time as the "Created Date"
    created_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row_data.append(created_date)
    
    ws.append(row_data)
    wb.save(EXCEL_FILE)

# Update Excel data
def update_excel(sheet_name, row_number, updated_data):
    wb = load_workbook(EXCEL_FILE)
    ws = wb[sheet_name]

    # Retrieve the current ID
    current_id = ws.cell(row=row_number, column=1).value  # Assuming ID is in the first column

    # Update the corresponding row in Excel (start from column 2)
    for idx, value in enumerate(updated_data):
        ws.cell(row=row_number, column=idx + 2, value=value)  # Start from column 2 to avoid overwriting ID

    # Write back the current ID to the first column
    ws.cell(row=row_number, column=1, value=current_id)
    
    wb.save(EXCEL_FILE)

# Delete row from Excel
def delete_row_in_excel(sheet_name, row_number):
    wb = load_workbook(EXCEL_FILE)
    ws = wb[sheet_name]
    ws.delete_rows(row_number)
    wb.save(EXCEL_FILE)

# Home Route
@app.route('/')
def home():
    return render_template('home.html')

@app.route('/report/')
def report():
    return render_template('report.html')

# AJAX endpoint to get relevant Master Data based on selection
@app.route('/get_master_data', methods=['POST'])
def get_master_data():
    selected_menu = request.form['menu']
    
    # Load data from Master Data
    if selected_menu == 'PIC':
        # Load data from Master PIC
        master_pic = read_excel('Master PIC')
        # Create a list of unique PIC entries (assuming PIC is in the first column)
        unique_entries = sorted(set([row[1] for row in master_pic]))
    else:
        # Load data from Master Data
        master_data = read_excel('Master Data')
        # Create a list of unique entries based on selected menu (column index)
        column_map = {
            'Function': 1,
            'Menu Utama': 2,
            'Sub Menu': 3,
            'Kode Produk': 4,
            'Produk': 5,
            'Nama Akad': 6
        }
        column_index = column_map[selected_menu]
        unique_entries = sorted(set([row[column_index] for row in master_data]))
    
    # Return as JSON for AJAX to handle
    return jsonify(unique_entries)

# AJAX endpoint to search the Input data based on selection
@app.route('/search_input', methods=['POST'])
def search_input():
    selected_menu = request.form['menu']
    selected_value = request.form['value']
    date_range = request.form.get('date_range', None)  # Get the date range if provided

    # Load data from Input sheet
    input_data = read_excel('Input')

    # Define mapping of columns to search
    column_map = {
        'Function': 1,
        'Menu Utama': 2,
        'Sub Menu': 3,
        'Kode Produk': 4,
        'Produk': 5,
        'Nama Akad': 6,
        'PIC': 13
    }
    column_index = column_map[selected_menu]

    # Filter input data based on the selected menu and value
    filtered_data = [row for row in input_data if row[column_index] == selected_value]

    # If date_range is provided, filter the data further by the Created Date (last column)
    if date_range:
        print(f"Received date range: {date_range}")
    
        # Case where there's no range, just a single date
        if 'to' not in date_range:
            try:
                datefilter = datetime.strptime(date_range, "%Y-%m-%d")
                print(f"Filtering for single date: {datefilter}")
                
                # Assuming row[-1] is a string in "%Y-%m-%d %H:%M:%S" format
                filtered_data = [
                    row for row in filtered_data
                    if datefilter.date() == datetime.strptime(row[-1], "%Y-%m-%d %H:%M:%S").date()
                ]
            except Exception as e:
                print(f"Error parsing single date: {e}")
        
        # Case where a date range is provided
        else:
            try:
                start_date_str, end_date_str = date_range.split(' to ')
                start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
                end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
                
                print(f"Filtering between {start_date} and {end_date}")
                
                # Filtering the data based on the date range
                filtered_data = [
                    row for row in filtered_data
                    if start_date <= datetime.strptime(row[-1], "%Y-%m-%d %H:%M:%S") <= end_date
                ]
            except Exception as e:
                print(f"Error parsing date range: {e}")

    return jsonify(filtered_data)

# Route to display Input Data
@app.route('/input_data/')
def show_input_data():
    data = read_excel('Input')
    return render_template('show_input_data.html', data=data, enumerate=enumerate)

# Route to add new Input Data
@app.route('/add_input_data/', methods=['GET', 'POST'])
def add_input_data():
    if request.method == 'POST':
        row_data = [
            request.form['fungsi'],
            request.form['MenuUtama'],
            request.form['SubMenu'],
            request.form['KodeProduk'],
            request.form['Produk'],
            request.form['NamaAkad'],
            int(request.form['scenario_positive']),
            int(request.form['scenario_negative']),
            int(request.form['scenario_positive']) + int(request.form['scenario_negative']),
            int(request.form['step_positive']),
            int(request.form['step_negative']),
            int(request.form['step_positive']) + int(request.form['step_negative']),
            request.form['PIC']
        ]
        print(row_data)
        write_to_excel('Input', row_data)
        return redirect(url_for('show_input_data'))

    master_data = read_excel('Master Data')
    master_pic = read_excel('Master PIC')
    return render_template('add_input_data.html', master_data=master_data, master_pic=master_pic)

# Route to edit Input Data
@app.route('/edit_input_data/<int:row_number>', methods=['GET', 'POST'])
def edit_input_data(row_number):
    if request.method == 'POST':
        updated_data = [
            request.form['fungsi'],
            request.form['MenuUtama'],
            request.form['SubMenu'],
            request.form['KodeProduk'],
            request.form['Produk'],
            request.form['NamaAkad'],
            int(request.form['scenario_positive']),
            int(request.form['scenario_negative']),
            int(request.form['scenario_positive']) + int(request.form['scenario_negative']),
            int(request.form['step_positive']),
            int(request.form['step_negative']),
            int(request.form['step_positive']) + int(request.form['step_negative']),
            request.form['PIC']
        ]
        print(updated_data)
        update_excel('Input', row_number, updated_data)
        return redirect(url_for('show_input_data'))

    data = read_excel('Input')
    master_data = read_excel('Master Data')
    master_pic = read_excel('Master PIC')
    return render_template('edit_input_data.html', data=data[row_number - 2], master_data=master_data, master_pic=master_pic)

# Route to delete Input Data
@app.route('/delete_input_data/<int:row_number>/')
def delete_input_data(row_number):
    delete_row_in_excel('Input', row_number)
    return redirect(url_for('show_input_data'))

#### MASTER DATA CRUD ####
# Route to display Master Data
@app.route('/master_data/')
def show_data():
    data = read_excel('Master Data')
    return render_template('show_data.html', data=data, enumerate=enumerate)

# Route to add new Master Data
@app.route('/add/', methods=['GET', 'POST'])
def add_data():
    if request.method == 'POST':
        row_data = [
            request.form['fungsi'],
            request.form['menu_utama'],
            request.form['sub_menu'],
            request.form['kode_produk'],
            request.form['nama_produk'],
            request.form['akad']
        ]
        write_to_excel('Master Data', row_data)
        return redirect(url_for('show_data'))
    return render_template('add_data.html')

# Route to edit a Master Data record
@app.route('/edit/<int:row_number>', methods=['GET', 'POST'])
def edit_data(row_number):
    if request.method == 'POST':
        updated_data = [
            request.form['fungsi'],
            request.form['menu_utama'],
            request.form['sub_menu'],
            request.form['kode_produk'],
            request.form['nama_produk'],
            request.form['akad']
        ]
        update_excel('Master Data', row_number, updated_data)
        return redirect(url_for('show_data'))

    data = read_excel('Master Data')
    return render_template('edit_data.html', data=data[row_number - 2])  # Adjust for header

# Route to delete a Master Data record
@app.route('/delete/<int:row_number>/')
def delete_data(row_number):
    delete_row_in_excel('Master Data', row_number)
    return redirect(url_for('show_data'))

#### MASTER PIC CRUD ####
# Route to display Master PIC
@app.route('/master_pic/')
def show_master_pic_data():
    data = read_excel('Master PIC')
    return render_template('show_master_pic_data.html', data=data, enumerate=enumerate)

# Route to add new Master PIC
@app.route('/add_master_pic/', methods=['GET', 'POST'])
def add_master_pic_data():
    if request.method == 'POST':
        row_data = [
            request.form['PIC'],
        ]
        write_to_excel('Master PIC', row_data)
        return redirect(url_for('show_master_pic_data'))
    return render_template('add_master_pic_data.html')

# Route to edit a Master PIC record
@app.route('/edit_master_pic/<int:row_number>', methods=['GET', 'POST'])
def edit_master_pic_data(row_number):
    if request.method == 'POST':
        updated_data = [
            request.form['PIC'],
        ]
        update_excel('Master PIC', row_number, updated_data)
        return redirect(url_for('show_master_pic_data'))

    data = read_excel('Master PIC')
    return render_template('edit_master_pic_data.html', data=data[row_number - 2])  # Adjust for header

# Route to delete a Master PIC record
@app.route('/delete_master_pic/<int:row_number>/')
def delete_master_pic_data(row_number):
    delete_row_in_excel('Master PIC', row_number)
    return redirect(url_for('show_master_pic_data'))

if __name__ == '__main__':
    app.run(debug=True)
